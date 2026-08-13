"""Excel exporter for the fixed seven-column qualified-job report."""

from __future__ import annotations

import json
import os
import re
import uuid
from contextlib import suppress
from pathlib import Path
from typing import Any
from urllib.parse import urlparse
from zipfile import BadZipFile, LargeZipFile

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import IllegalCharacterError, InvalidFileException

from .db import Database, utc_now


EXCEL_COLUMNS = [
    "岗位名称", "城市", "薪资范围", "岗位职责", "任职要求",
    "加分项", "岗位详情链接",
]
REVIEW_COLUMNS = ["岗位名称", "城市", "薪资范围", "复核理由", "置信度", "岗位详情链接"]
EXPORT_FAILURE_EXCEPTIONS = (
    KeyError, OSError, RuntimeError, TypeError, ValueError,
    InvalidFileException, IllegalCharacterError, BadZipFile, LargeZipFile,
)

_INVALID_FILENAME_CHARS = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
_SPREADSHEET_FORMULA_PREFIXES = ("=", "+", "-", "@")


def _require_snapshot_ready(run: dict[str, Any]) -> None:
    if run["status"] == "running":
        raise RuntimeError(
            f"Run {run['run_id']} is still running and cannot be frozen or exported"
        )


def sanitize_filename_component(value: str, fallback: str = "未指定") -> str:
    cleaned = _INVALID_FILENAME_CHARS.sub("_", str(value or "").strip())
    cleaned = re.sub(r"\s+", "_", cleaned).strip(" ._")
    return (cleaned or fallback)[:80]


def _safe_spreadsheet_value(value: Any) -> Any:
    if not isinstance(value, str):
        return value
    if value.lstrip(" \t\r\n").startswith(_SPREADSHEET_FORMULA_PREFIXES):
        return f"'{value}"
    return value


def _is_safe_job_url(value: Any) -> bool:
    parsed = urlparse(str(value or "").strip())
    return parsed.scheme == "https" and parsed.hostname == "www.zhipin.com"


def _safe_spreadsheet_rows(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {column: _safe_spreadsheet_value(value) for column, value in row.items()}
        for row in rows
    ]


def _display_list(value: Any, fallback: str) -> str:
    if value in (None, ""):
        return fallback
    parsed = value
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in ("待处理", "未注明"):
            return stripped
        try:
            parsed = json.loads(stripped)
        except (json.JSONDecodeError, TypeError, ValueError):
            return stripped or fallback
    if isinstance(parsed, list):
        items = [str(item).strip() for item in parsed if str(item).strip()]
        if not items:
            return fallback
        if items == ["无"]:
            return "无"
        return "\n".join(
            f"{index}. {item}" for index, item in enumerate(items, 1)
        )
    return str(parsed).strip() or fallback


def _deduplicate_jobs(jobs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    ranked_status = {"completed": 2, "manual_review": 1}
    groups: list[dict[str, Any]] = []
    for job in jobs:
        stable_id = str(job.get("source_job_id") or "").strip()
        normalized_url = str(job.get("normalized_job_url") or "").strip()
        sources = {stable_id} if stable_id else {
            f"row:{job.get('task_id', '')}:{job.get('job_id', '')}",
        }
        urls = {normalized_url} if normalized_url else set()
        matches = [
            index for index, group in enumerate(groups)
            if group["sources"] & sources or group["urls"] & urls
        ]
        if not matches:
            groups.append({"sources": sources, "urls": urls, "job": job})
            continue
        group = groups[matches[0]]
        candidates = [group["job"], job]
        group["sources"].update(sources)
        group["urls"].update(urls)
        for index in reversed(matches[1:]):
            merged = groups.pop(index)
            group["sources"].update(merged["sources"])
            group["urls"].update(merged["urls"])
            candidates.append(merged["job"])
        group["job"] = max(
            candidates,
            key=lambda item: ranked_status.get(str(item.get("ai_status")), 0),
        )
    return [group["job"] for group in groups]


def _available_output_path(path: Path) -> Path:
    if not path.exists():
        return path
    index = 2
    while True:
        candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def _write_workbook(
    path: Path,
    rows: list[dict[str, Any]],
    review_rows: list[dict[str, Any]],
) -> None:
    dataframe = pd.DataFrame(_safe_spreadsheet_rows(rows), columns=EXCEL_COLUMNS)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="岗位信息")
        sheet = writer.book["岗位信息"]
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for index, width in enumerate((24, 14, 16, 48, 48, 40, 48), 1):
            sheet.column_dimensions[get_column_letter(index)].width = width
        for row in sheet.iter_rows(min_row=2):
            for index in (4, 5, 6):
                row[index - 1].alignment = Alignment(
                    wrap_text=True, vertical="top",
                )
            if _is_safe_job_url(row[6].value):
                row[6].hyperlink = str(row[6].value)
                row[6].style = "Hyperlink"
        if review_rows:
            review_frame = pd.DataFrame(
                _safe_spreadsheet_rows(review_rows), columns=REVIEW_COLUMNS,
            )
            review_frame.to_excel(
                writer, index=False, sheet_name="待人工确认",
            )
            review_sheet = writer.book["待人工确认"]
            review_sheet.freeze_panes = "A2"
            review_sheet.auto_filter.ref = review_sheet.dimensions
            for index, width in enumerate((24, 14, 16, 48, 12, 48), 1):
                review_sheet.column_dimensions[get_column_letter(index)].width = width
            for row in review_sheet.iter_rows(min_row=2):
                row[3].alignment = Alignment(
                    wrap_text=True, vertical="top",
                )
                if _is_safe_job_url(row[5].value):
                    row[5].hyperlink = str(row[5].value)
                    row[5].style = "Hyperlink"


def _qualified_rows(jobs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "岗位名称": job["job_name"] or "未注明",
            "城市": job["city"] or "未注明",
            "薪资范围": job["salary_range"] or "未注明",
            "岗位职责": _display_list(
                job["job_responsibilities"], "未注明",
            ),
            "任职要求": _display_list(
                job["job_requirements"], "未注明",
            ),
            "加分项": _display_list(job["bonus_points"], "未注明"),
            "岗位详情链接": job["job_url"],
        }
        for job in jobs
        if job["crawl_status"] == "completed"
        and job["ai_status"] == "completed"
        and job.get("availability_status", "available")
            != "confirmed_unavailable"
    ]


def _review_rows(jobs: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [
        {
            "岗位名称": job["job_name"] or "未注明",
            "城市": job["city"] or "未注明",
            "薪资范围": job["salary_range"] or "未注明",
            "复核理由": job["relevance_reason"] or "未注明",
            "置信度": job["relevance_confidence"],
            "岗位详情链接": job["job_url"],
        }
        for job in jobs
        if job["crawl_status"] == "completed"
        and job["ai_status"] == "manual_review"
        and job.get("availability_status", "available")
            != "confirmed_unavailable"
    ]


def _verify_workbook(path: Path, expected_rows: int) -> None:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        if "岗位信息" not in workbook.sheetnames:
            raise ValueError("导出缺少岗位信息工作表")
        sheet = workbook["岗位信息"]
        headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
        if headers != EXCEL_COLUMNS:
            raise ValueError(f"导出表头错误: {headers}")
        if sheet.max_row - 1 != expected_rows:
            raise ValueError(
                f"导出行数错误: {sheet.max_row - 1} != {expected_rows}"
            )
    finally:
        workbook.close()


def freeze_strategy_run_snapshot(
    database: Database,
    strategy_id: str,
    run_id: str,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    run = database.get_run(run_id)
    if run is None or run["strategy_id"] != strategy_id:
        raise KeyError(run_id)
    _require_snapshot_ready(run)
    if run["export_rows_json"]:
        return (
            json.loads(run["export_rows_json"]),
            json.loads(run["review_rows_json"] or "[]"),
        )
    jobs = database.list_strategy_jobs_as_of_run(strategy_id, run_id)
    rows = _qualified_rows(jobs)
    review_rows = _review_rows(jobs)
    database.update_run(
        run_id,
        export_rows_json=json.dumps(rows, ensure_ascii=False),
        review_rows_json=json.dumps(review_rows, ensure_ascii=False),
        export_snapshot_at=database.get_run(run_id)["finished_at"] or utc_now(),
    )
    return rows, review_rows


def export_strategy_run(
    database: Database,
    strategy_id: str,
    run_id: str,
    output_dir: str | Path | None = None,
) -> Path:
    strategy = database.get_strategy(strategy_id)
    run = database.get_run(run_id)
    if strategy is None:
        raise KeyError(strategy_id)
    if run is None or run["strategy_id"] != strategy_id:
        raise KeyError(run_id)
    _require_snapshot_ready(run)
    current = Path(run["output_path"]).expanduser() if run["output_path"] else None
    if (
        run["export_status"] == "completed"
        and current is not None
        and current.is_file()
    ):
        latest_run = database.get_latest_run(strategy_id)
        if latest_run is not None and latest_run["run_id"] == run_id:
            database.update_strategy(
                strategy_id, latest_output_path=str(current),
            )
        return current
    rows, review_rows = freeze_strategy_run_snapshot(
        database, strategy_id, run_id,
    )
    temporary_path: Path | None = None
    try:
        target_dir = Path(
            output_dir
            or Path.home() / ".boss-zhipin-job-research" / "job-result"
        ).expanduser()
        target_dir.mkdir(parents=True, exist_ok=True)
        filename = (
            f"BOSS直聘_{sanitize_filename_component(strategy['target_role'])}_"
            f"Run{int(run['run_number']):03d}_累计{len(rows)}条.xlsx"
        )
        output_path = _available_output_path(target_dir / filename)
        temporary_path = output_path.with_name(
            f".{output_path.stem}.{uuid.uuid4().hex}.tmp.xlsx"
        )
        _write_workbook(temporary_path, rows, review_rows)
        _verify_workbook(temporary_path, len(rows))
        os.replace(temporary_path, output_path)
        temporary_path = None
        database.update_run(
            run_id,
            export_status="completed",
            output_path=str(output_path),
            cumulative_export_count=len(rows),
            export_error="",
        )
        latest_run = database.get_latest_run(strategy_id)
        if latest_run is not None and latest_run["run_id"] == run_id:
            database.update_strategy(
                strategy_id, latest_output_path=str(output_path),
            )
        return output_path
    except EXPORT_FAILURE_EXCEPTIONS as exc:
        database.update_run(
            run_id,
            export_status="failed",
            export_error=str(exc),
        )
        raise
    finally:
        if temporary_path is not None:
            with suppress(OSError):
                temporary_path.unlink(missing_ok=True)


def export_task_to_excel(
    database: Database,
    task_id: str,
    output_dir: str | Path | None = None,
) -> Path:
    return export_tasks_to_excel(database, [task_id], output_dir)


def export_tasks_to_excel(
    database: Database,
    task_ids: list[str],
    output_dir: str | Path | None = None,
) -> Path:
    if not task_ids:
        raise ValueError("至少需要一个任务")
    tasks = []
    all_jobs = []
    for task_id in task_ids:
        task = database.get_task(task_id)
        if task is None:
            raise KeyError(task_id)
        tasks.append(task)
        all_jobs.extend(database.list_jobs(task_id))
    task = tasks[0]
    all_jobs = _deduplicate_jobs(all_jobs)
    jobs = [
        job for job in all_jobs
        if job["crawl_status"] == "completed" and job["ai_status"] == "completed"
    ]
    rows = []
    for job in jobs:
        pending = "待处理" if job["ai_status"] != "completed" else "未注明"
        rows.append(
            {
                "岗位名称": job["job_name"] or "未注明",
                "城市": job["city"] or "未注明",
                "薪资范围": job["salary_range"] or "未注明",
                "岗位职责": _display_list(job["job_responsibilities"], pending),
                "任职要求": _display_list(job["job_requirements"], pending),
                "加分项": _display_list(job["bonus_points"], pending),
                "岗位详情链接": job["job_url"],
            }
        )
    review_rows = [
        {
            "岗位名称": job["job_name"] or "未注明",
            "城市": job["city"] or "未注明",
            "薪资范围": job["salary_range"] or "未注明",
            "复核理由": job["relevance_reason"] or "未注明",
            "置信度": job["relevance_confidence"],
            "岗位详情链接": job["job_url"],
        }
        for job in all_jobs
        if job["crawl_status"] == "completed" and job["ai_status"] == "manual_review"
    ]
    target_dir = Path(output_dir or Path.home() / ".boss-zhipin-job-research" / "job-result").expanduser()
    target_dir.mkdir(parents=True, exist_ok=True)
    city_name = "_".join(dict.fromkeys(item["city"] for item in tasks))
    filename = (
        f"BOSS直聘_{sanitize_filename_component(city_name)}_"
        f"{sanitize_filename_component(task['keyword'])}_岗位信息.xlsx"
    )
    output_path = _available_output_path(target_dir / filename)
    temporary_path = output_path.with_name(f".{output_path.stem}.{uuid.uuid4().hex}.tmp.xlsx")
    try:
        _write_workbook(temporary_path, rows, review_rows)
        os.replace(temporary_path, output_path)
    finally:
        temporary_path.unlink(missing_ok=True)
    for task_id in task_ids:
        database.update_task(task_id, output_path=str(output_path))
    return output_path
