import json
import os
import tempfile
import threading
import unittest
from unittest import mock
from datetime import datetime, timedelta, timezone
from pathlib import Path

from boss_app import db as db_module
from boss_app.db import Database, uses_qualified_target, utc_now
from boss_app.ai_parser import (
    AIConfig, AIParseError, JDParser, ParsedJD, validate_payload,
    validate_target_payload,
)
from boss_app.exporter import EXCEL_COLUMNS, export_task_to_excel, export_tasks_to_excel
from boss_app.job_type_parser import normalize_job_type
from boss_app.login_manager import LoginManager, LoginState, LoginStatus
from boss_app.salary_parser import parse_salary
from boss_app.task_manager import TaskManager
from boss_app.strategy_model import StrategySpec
from boss_app.collector import Collector
from scripts import boss_cdp_raw as core
from openpyxl import load_workbook


class RunModeTests(unittest.TestCase):
    def test_run_modes_resolve_fixed_and_custom_limits(self):
        from boss_app import task_manager as task_manager_module

        resolver = getattr(task_manager_module, "resolve_job_limit", None)
        limits = getattr(task_manager_module, "RUN_MODE_LIMITS", None)
        self.assertIsNotNone(resolver, "缺少分阶段运行数量解析器")
        self.assertEqual(
            limits,
            {
                "10条验证": 10,
                "20条稳定性测试": 20,
                "50条扩容测试": 50,
                "100条批量测试": 100,
                "自定义数量": None,
            },
        )
        self.assertEqual(resolver("10条验证", 999), 10)
        self.assertEqual(resolver("100条批量测试", 1), 100)
        self.assertEqual(resolver("自定义数量", 137), 137)
        with self.assertRaises(ValueError):
            resolver("自定义数量", 0)


class SalaryParserTests(unittest.TestCase):
    def test_salary_months_are_split_without_converting_units(self):
        parts = parse_salary("20-30K·14薪")
        self.assertEqual(parts.raw, "20-30K·14薪")
        self.assertEqual(parts.salary_range, "20-30K")
        self.assertEqual(parts.salary_months, "14薪")

    def test_daily_salary_is_preserved(self):
        parts = parse_salary("200-300元/天")
        self.assertEqual(parts.salary_range, "200-300元/天")
        self.assertEqual(parts.salary_months, "未注明")


class JobTypeParserTests(unittest.TestCase):
    def test_explicit_page_signal_precedes_ai(self):
        self.assertEqual(normalize_job_type(title="AI运营实习生", ai_value="全职"), "实习")

    def test_fresh_graduate_phrase_does_not_imply_campus_hiring(self):
        self.assertEqual(normalize_job_type(jd="应届生可投"), "未注明")

    def test_ai_value_is_limited_to_fixed_enum(self):
        self.assertEqual(normalize_job_type(ai_value="全职"), "全职")
        self.assertEqual(normalize_job_type(ai_value="自由职业"), "未注明")


class DatabaseTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.db = Database(Path(self.tempdir.name) / "jobs.db")
        self.task_id = self.db.create_task("AI运营", "深圳", max_pages=1, max_jobs=10)

    def tearDown(self):
        self.tempdir.cleanup()

    def test_list_page_cursor_defaults_and_updates(self):
        self.assertEqual(self.db.get_task(self.task_id).get("list_next_page"), 1)

        self.db.update_task(self.task_id, list_next_page=6)

        self.assertEqual(self.db.get_task(self.task_id)["list_next_page"], 6)

    def test_job_id_is_deduplicated_within_task(self):
        job = {"job_id": "abc", "title": "AI运营", "job_link": "https://example.test/abc"}
        self.assertTrue(self.db.upsert_job(self.task_id, job))
        self.assertFalse(self.db.upsert_job(self.task_id, job))
        self.assertEqual(len(self.db.list_jobs(self.task_id)), 1)

    def test_normalized_job_link_deduplicates_different_job_ids(self):
        first = {
            "job_id": "first",
            "title": "AI运营",
            "job_link": "HTTPS://WWW.ZHIPIN.COM/job_detail/abc123.html?lid=one#top",
        }
        duplicate = {
            "job_id": "second",
            "title": "AI运营（重复链接）",
            "job_link": "https://www.zhipin.com/job_detail/abc123.html/",
        }
        self.assertTrue(self.db.upsert_job(self.task_id, first))
        self.assertFalse(self.db.upsert_job(self.task_id, duplicate))
        rows = self.db.list_jobs(self.task_id)
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["job_id"], "first")
        self.assertEqual(
            rows[0]["normalized_job_url"],
            "https://www.zhipin.com/job_detail/abc123.html",
        )

    def test_same_job_can_belong_to_different_tasks(self):
        other_task = self.db.create_task("AI产品", "深圳")
        job = {"job_id": "abc", "title": "AI运营"}
        self.assertTrue(self.db.upsert_job(self.task_id, job))
        self.assertTrue(self.db.upsert_job(other_task, job))

    def test_stale_processing_state_is_recovered(self):
        job = {"job_id": "abc", "title": "AI运营"}
        self.db.upsert_job(self.task_id, job)
        token = "stale-token"
        self.assertTrue(self.db.reserve_worker(self.task_id, token))
        self.db.update_job(self.task_id, "abc", crawl_status="processing", ai_status="processing")
        old = (datetime.now(timezone.utc) - timedelta(minutes=5)).isoformat()
        with self.db.connect() as connection:
            connection.execute(
                "UPDATE tasks SET worker_heartbeat_at=? WHERE task_id=?", (old, self.task_id)
            )
        self.assertEqual(self.db.recover_interrupted(stale_seconds=60), [self.task_id])
        task = self.db.get_task(self.task_id)
        row = self.db.list_jobs(self.task_id)[0]
        self.assertEqual(task["status"], "pending")
        self.assertIsNone(task["worker_token"])
        self.assertEqual(row["crawl_status"], "pending")
        self.assertEqual(row["ai_status"], "pending")

    def test_worker_lease_allows_only_one_active_task(self):
        other_task = self.db.create_task("产品运营", "深圳")
        self.assertTrue(self.db.reserve_worker(self.task_id, "one"))
        self.assertFalse(self.db.reserve_worker(other_task, "two"))
        self.db.release_worker(self.task_id, "one")
        self.assertTrue(self.db.reserve_worker(other_task, "two"))

    def test_orphan_processing_without_worker_is_recovered(self):
        self.db.upsert_job(self.task_id, {"job_id": "orphan", "title": "AI运营"})
        self.db.update_task(self.task_id, status="processing")
        self.db.update_job(
            self.task_id, "orphan", crawl_status="processing", ai_status="processing",
        )

        recovered = self.db.recover_interrupted(stale_seconds=60)

        self.assertEqual(recovered, [self.task_id])
        self.assertEqual(self.db.get_task(self.task_id)["status"], "pending")
        row = self.db.list_jobs(self.task_id)[0]
        self.assertEqual(row["crawl_status"], "pending")
        self.assertEqual(row["ai_status"], "pending")

    def test_existing_ten_jobs_count_toward_expansion_to_fifty(self):
        for index in range(10):
            self.db.upsert_job(
                self.task_id,
                {
                    "job_id": f"job-{index}",
                    "title": f"AI运营 {index}",
                    "job_link": f"https://example.test/job/{index}",
                },
            )
            self.db.update_job(self.task_id, f"job-{index}", crawl_status="completed")
        updater = getattr(self.db, "update_task_limits", None)
        self.assertIsNotNone(updater, "缺少历史任务扩容接口")

        updater(self.task_id, max_jobs=50, max_pages=3, run_mode="50条扩容测试")

        task = self.db.get_task(self.task_id)
        self.assertEqual(task["max_jobs"], 50)
        self.assertEqual(task["max_pages"], 3)
        self.assertEqual(task["run_mode"], "50条扩容测试")
        self.assertEqual(len(self.db.list_jobs(self.task_id)), 10)

    def test_national_task_expansion_uses_qualified_count_as_minimum(self):
        task_id = self.db.create_task("AI运营", "全国", max_jobs=50)
        for index in range(60):
            job_id = f"national-{index}"
            self.db.upsert_job(task_id, {"job_id": job_id, "title": f"候选 {index}"})
            self.db.update_job(
                task_id,
                job_id,
                crawl_status="completed",
                ai_status="completed" if index < 50 else "irrelevant",
            )

        self.db.update_task_limits(
            task_id, max_jobs=55, max_pages=10, run_mode="自定义数量",
        )

        task = self.db.get_task(task_id)
        self.assertEqual(task["max_jobs"], 55)
        self.assertEqual(len(self.db.list_jobs(task_id)), 60)

    def test_schema_migration_preserves_existing_jobs(self):
        for index in range(10):
            self.db.upsert_job(
                self.task_id,
                {
                    "job_id": f"saved-{index}",
                    "title": f"已保存岗位 {index}",
                    "job_link": f"https://example.test/saved/{index}",
                },
            )
        with self.db.connect() as connection:
            connection.execute("DROP INDEX idx_jobs_url")
            connection.execute("ALTER TABLE jobs DROP COLUMN normalized_job_url")
            connection.execute("ALTER TABLE tasks DROP COLUMN run_mode")
            task_columns = {
                row["name"] for row in connection.execute("PRAGMA table_info(tasks)")
            }
            if "list_next_page" in task_columns:
                connection.execute("ALTER TABLE tasks DROP COLUMN list_next_page")

        migrated = Database(self.db.path)

        self.assertEqual(len(migrated.list_jobs(self.task_id)), 10)
        self.assertTrue(all(row["normalized_job_url"] for row in migrated.list_jobs(self.task_id)))
        self.assertEqual(migrated.get_task(self.task_id)["run_mode"], "10条验证")
        self.assertEqual(migrated.get_task(self.task_id)["list_next_page"], 1)

    def test_city_name_is_preferred_and_normalized(self):
        self.db.upsert_job(
            self.task_id,
            {
                "job_id": "city",
                "title": "AI运营",
                "city_name": "深圳",
                "location": "深圳·南山区·科技园",
            },
        )

        self.assertEqual(self.db.list_jobs(self.task_id)[0]["city"], "深圳")

    def test_snapshot_counts_only_completed_detail_and_ai_as_qualified(self):
        for job_id in ("qualified", "irrelevant", "pending"):
            self.db.upsert_job(self.task_id, {"job_id": job_id, "title": job_id})
        self.db.update_job(
            self.task_id, "qualified", crawl_status="completed", ai_status="completed",
        )
        self.db.update_job(
            self.task_id, "irrelevant", crawl_status="completed", ai_status="irrelevant",
        )

        snapshot = self.db.snapshot(self.task_id)

        self.assertEqual(snapshot["qualified"], 1)
        self.assertEqual(snapshot["irrelevant"], 1)
        self.assertEqual(snapshot["failed"], 0)

    def test_city_ai_product_operations_uses_relevance_filter_only(self):
        predicate = getattr(db_module, "uses_ai_relevance_filter", None)
        self.assertIsNotNone(predicate, "缺少城市 AI 产品运营相关性判定")
        self.assertTrue(predicate("AI产品运营", "北京"))
        self.assertTrue(predicate(" AI 产品运营 ", "上海"))
        self.assertFalse(predicate("AI运营", "深圳"))
        self.assertFalse(uses_qualified_target("AI产品运营", "北京"))


class _FakeResponse:
    def __init__(self, payload, status_error=None):
        self.payload = payload
        self.status_error = status_error

    def raise_for_status(self):
        if self.status_error:
            raise self.status_error

    def json(self):
        return self.payload


class _FakeSession:
    def __init__(self, responses):
        self.responses = list(responses)
        self.calls = []

    def post(self, url, **kwargs):
        self.calls.append((url, kwargs))
        response = self.responses.pop(0)
        if isinstance(response, BaseException):
            raise response
        return response


class AIParserTests(unittest.TestCase):
    def setUp(self):
        self.config = AIConfig(
            api_key="test-key", base_url="https://ai.example.test/v1",
            model="test-model", timeout=3, max_retries=1,
        )

    def test_ai_json_schema_is_strictly_validated(self):
        payload = {
            "is_ai_operations": True,
            "job_type": "全职",
            "job_responsibilities": ["负责 AI 产品运营"],
            "job_requirements": ["具备运营经验"],
            "bonus_points": [],
        }
        parsed = validate_payload(payload)
        self.assertTrue(parsed.is_ai_operations)
        self.assertEqual(parsed.job_type, "全职")
        self.assertEqual(parsed.bonus_points, ["无"])
        with self.assertRaises(AIParseError):
            validate_payload({"job_type": "全职"})
        for invalid in ("true", 1, None):
            with self.subTest(invalid=invalid):
                with self.assertRaises(AIParseError):
                    validate_payload(dict(payload, is_ai_operations=invalid))

    def test_empty_jd_never_calls_ai(self):
        session = _FakeSession([])
        parser = JDParser(self.config, session=session, sleep_fn=lambda _seconds: None)
        with self.assertRaisesRegex(ValueError, "完整 JD 为空"):
            parser.parse("")
        self.assertEqual(session.calls, [])

    def test_ai_failure_retries_then_raises_normalized_error(self):
        session = _FakeSession([RuntimeError("temporary"), RuntimeError("still down")])
        parser = JDParser(self.config, session=session, sleep_fn=lambda _seconds: None)
        with self.assertRaises(AIParseError):
            parser.parse("职位描述：负责 AI 运营")
        self.assertEqual(len(session.calls), 2)

    def test_valid_ai_response_preserves_explicit_job_type(self):
        content = json.dumps(
            {
                "is_ai_operations": True,
                "job_type": "全职",
                "job_responsibilities": ["负责 AI 产品运营"],
                "job_requirements": ["本科"],
                "bonus_points": ["有 AIGC 经验者优先"],
            },
            ensure_ascii=False,
        )
        session = _FakeSession([_FakeResponse({"choices": [{"message": {"content": content}}]})])
        parser = JDParser(self.config, session=session, sleep_fn=lambda _seconds: None)
        parsed = parser.parse("完整 JD", {"job_name": "AI运营实习生"})
        self.assertEqual(parsed.job_type, "实习")
        self.assertEqual(session.calls[0][0], "https://ai.example.test/v1/chat/completions")

    def test_target_payload_supports_match_review_and_irrelevant(self):
        base = {
            "match_status": "matched",
            "role_category": "新媒体运营",
            "relevance_reason": "核心职责是新媒体账号运营",
            "relevance_confidence": 0.93,
            "job_type": "全职",
            "job_responsibilities": ["运营新媒体账号"],
            "job_requirements": ["具备内容运营经验"],
            "bonus_points": ["有短视频经验优先"],
        }
        for status in ("matched", "manual_review", "irrelevant"):
            with self.subTest(status=status):
                parsed = validate_target_payload(dict(base, match_status=status))
                self.assertEqual(parsed.match_status, status)
        for invalid in (-0.1, 1.1, True, "0.9"):
            with self.subTest(confidence=invalid), self.assertRaises(AIParseError):
                validate_target_payload(dict(base, relevance_confidence=invalid))
        with self.assertRaises(AIParseError):
            validate_target_payload(dict(base, relevance_reason=""))

    def test_target_parser_sends_confirmed_target_without_expanding_it(self):
        content = json.dumps(
            {
                "match_status": "matched",
                "role_category": "新媒体运营",
                "relevance_reason": "JD核心职责与目标一致",
                "relevance_confidence": 0.95,
                "job_type": "全职",
                "job_responsibilities": ["负责公众号运营"],
                "job_requirements": ["两年运营经验"],
                "bonus_points": ["无"],
            },
            ensure_ascii=False,
        )
        session = _FakeSession([_FakeResponse({"choices": [{"message": {"content": content}}]})])
        parser = JDParser(self.config, session=session, sleep_fn=lambda _seconds: None)

        parsed = parser.parse_target(
            "完整 JD", target_role="新媒体运营", target_type="exact_role",
            context={"job_name": "新媒体运营专员"},
        )

        request = session.calls[0][1]["json"]
        user_payload = json.loads(request["messages"][1]["content"])
        self.assertEqual(user_payload["用户确认目标"], "新媒体运营")
        self.assertEqual(user_payload["目标类型"], "exact_role")
        self.assertNotIn("短视频运营", request["messages"][0]["content"])
        self.assertEqual(parsed.match_status, "matched")


class GenericTargetDatabaseTests(unittest.TestCase):
    def test_target_fields_and_manual_review_are_additive(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI产品", "北京", target_role="AI产品相关岗位",
                target_type="domain_scope",
            )
            database.upsert_job(
                task_id,
                {"job_id": "review", "title": "智能产品岗位", "job_link": "https://example.test/review"},
            )
            database.update_job(
                task_id, "review", crawl_status="completed", ai_status="manual_review",
                role_category="产品研发", relevance_reason="JD没有说明是否参与产品职责",
                relevance_confidence=0.52,
            )

            task = database.get_task(task_id)
            job = database.list_jobs(task_id)[0]
            snapshot = database.snapshot(task_id)
            self.assertEqual(task["target_role"], "AI产品相关岗位")
            self.assertEqual(task["target_type"], "domain_scope")
            self.assertEqual(job["ai_status"], "manual_review")
            self.assertEqual(job["role_category"], "产品研发")
            self.assertEqual(job["relevance_confidence"], 0.52)
            self.assertEqual(snapshot["manual_review"], 1)
            self.assertEqual(database.next_jobs(task_id, "ai"), [])


class ExcelExporterTests(unittest.TestCase):
    def test_multi_task_export_merges_matched_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_ids = []
            for city in ("北京", "上海"):
                task_id = database.create_task("新媒体运营", city)
                task_ids.append(task_id)
                database.upsert_job(
                    task_id,
                    {
                        "job_id": city, "title": f"{city}岗位", "city_name": city,
                        "job_link": f"https://example.test/{city}",
                    },
                )
                database.update_job(task_id, city, crawl_status="completed", ai_status="completed")

            output = export_tasks_to_excel(database, task_ids, directory)
            sheet = load_workbook(output)["岗位信息"]

            self.assertEqual((sheet.max_row, sheet.max_column), (3, 7))
            self.assertEqual([sheet["A2"].value, sheet["A3"].value], ["北京岗位", "上海岗位"])

    def test_multi_task_export_deduplicates_the_same_source_job(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_ids = []
            identities = (
                ("北京", "shared-source", "https://example.test/job/one"),
                ("上海", "shared-source", "https://example.test/job/two"),
                ("广州", "different-source", "https://example.test/job/two"),
            )
            for city, source_id, job_url in identities:
                task_id = database.create_task("新媒体运营", city)
                task_ids.append(task_id)
                job_id = f"duplicate-{city}"
                database.upsert_job(
                    task_id,
                    {
                        "job_id": job_id, "encrypt_job_id": source_id,
                        "title": "同一岗位", "city_name": city,
                        "job_link": job_url,
                    },
                )
                database.update_job(
                    task_id, job_id, crawl_status="completed", ai_status="completed",
                )

            sheet = load_workbook(
                export_tasks_to_excel(database, task_ids, directory),
            )["岗位信息"]

            self.assertEqual(sheet.max_row - 1, 1)

    def test_excel_has_fixed_columns_and_clickable_hyperlink(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            database.upsert_job(
                task_id,
                {
                    "job_id": "abc", "title": "AI运营", "city_name": "深圳",
                    "location": "深圳·南山区",
                    "salary": "20-30K·14薪",
                    "job_link": "https://www.zhipin.com/job_detail/abc.html",
                    "experience": "1-3年", "education": "本科",
                },
            )
            database.update_job(
                task_id, "abc", salary_range="20-30K", salary_months="14薪",
                job_type="全职", crawl_status="completed", ai_status="completed",
                job_responsibilities=["负责内容运营", "分析运营数据"],
                job_requirements=["本科"], bonus_points=["无"], full_jd="原始完整 JD",
            )
            output = export_task_to_excel(database, task_id, directory)
            workbook = load_workbook(output)
            sheet = workbook["岗位信息"]
            self.assertEqual([cell.value for cell in sheet[1]], EXCEL_COLUMNS)
            self.assertEqual(EXCEL_COLUMNS, [
                "岗位名称", "城市", "薪资范围", "岗位职责", "任职要求",
                "加分项", "岗位详情链接",
            ])
            self.assertEqual((sheet.max_row, sheet.max_column), (2, 7))
            self.assertEqual(sheet.freeze_panes, "A2")
            self.assertEqual(sheet["B2"].value, "深圳")
            self.assertEqual(sheet["D2"].value, "1. 负责内容运营\n2. 分析运营数据")
            self.assertEqual(sheet["F2"].value, "无")
            self.assertEqual(
                sheet["G2"].hyperlink.target,
                "https://www.zhipin.com/job_detail/abc.html",
            )
            self.assertEqual(database.list_jobs(task_id)[0]["full_jd"], "原始完整 JD")

    def test_excel_escapes_formula_text_and_rejects_untrusted_hyperlinks(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            database.upsert_job(
                task_id,
                {
                    "job_id": "unsafe", "title": "=1+1", "city_name": "+深圳",
                    "salary": "-1+1", "job_link": "https://evil.test/job/unsafe",
                },
            )
            database.update_job(
                task_id, "unsafe", crawl_status="completed", ai_status="completed",
                salary_range="-1+1",
                job_responsibilities="=HYPERLINK(\"https://evil.test\")",
                job_requirements="@SUM(1,1)", bonus_points="\t=1+1",
            )

            workbook = load_workbook(export_task_to_excel(database, task_id, directory))
            sheet = workbook["岗位信息"]

            self.assertEqual(sheet["A2"].value, "'=1+1")
            self.assertEqual(sheet["B2"].value, "'+深圳")
            self.assertEqual(sheet["C2"].value, "'-1+1")
            self.assertEqual(sheet["D2"].value, "'=HYPERLINK(\"https://evil.test\")")
            self.assertEqual(sheet["E2"].value, "'@SUM(1,1)")
            self.assertEqual(sheet["F2"].value, "'=1+1")
            self.assertIsNone(sheet["G2"].hyperlink)

    def test_excel_regeneration_preserves_the_previous_file_atomically(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            database.upsert_job(
                task_id,
                {"job_id": "one", "title": "岗位一", "job_link": "https://example.test/one"},
            )
            database.update_job(
                task_id, "one", crawl_status="completed", ai_status="completed",
            )
            first = export_task_to_excel(database, task_id, directory)
            self.assertEqual(load_workbook(first)["岗位信息"].max_row - 1, 1)
            database.upsert_job(
                task_id,
                {"job_id": "two", "title": "岗位二", "job_link": "https://example.test/two"},
            )
            database.update_job(
                task_id, "two", crawl_status="completed", ai_status="completed",
            )

            with mock.patch("boss_app.exporter.os.replace", wraps=os.replace) as replace:
                second = export_task_to_excel(database, task_id, directory)

            self.assertNotEqual(second, first)
            self.assertEqual(load_workbook(first)["岗位信息"].max_row - 1, 1)
            self.assertEqual(load_workbook(second)["岗位信息"].max_row - 1, 2)
            replace.assert_called_once()
            self.assertEqual(list(Path(directory).glob("*.tmp.xlsx")), [])

    def test_excel_exports_only_qualified_jobs(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "全国", max_jobs=50)
            for index in range(60):
                job_id = f"job-{index}"
                database.upsert_job(
                    task_id,
                    {
                        "job_id": job_id,
                        "title": f"AI运营 {index}",
                        "city_name": "上海",
                        "job_link": f"https://example.test/{job_id}",
                    },
                )
                database.update_job(
                    task_id, job_id, crawl_status="completed",
                    ai_status="completed" if index < 50 else "irrelevant",
                    job_responsibilities=["负责 AI 运营"],
                    job_requirements=["具备运营经验"], bonus_points=["无"],
                )

            output = export_task_to_excel(database, task_id, directory)
            sheet = load_workbook(output)["岗位信息"]

            self.assertEqual((sheet.max_row, sheet.max_column), (51, 7))
            self.assertEqual(sheet["A2"].value, "AI运营 0")
            self.assertEqual(sheet["A51"].value, "AI运营 49")

    def test_city_export_keeps_irrelevant_rows_for_audit_only(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI产品运营", "北京")
            for job_id in ("qualified", "irrelevant"):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": job_id, "title": job_id,
                        "job_link": f"https://example.test/{job_id}",
                    },
                )
                database.update_job(
                    task_id, job_id, crawl_status="completed",
                    ai_status="completed" if job_id == "qualified" else "irrelevant",
                )

            output = export_task_to_excel(database, task_id, directory)
            sheet = load_workbook(output)["岗位信息"]

            self.assertEqual((sheet.max_row, sheet.max_column), (2, 7))
            self.assertEqual(sheet["A2"].value, "qualified")
            rows = database.list_jobs(task_id)
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[1]["ai_status"], "irrelevant")

    def test_excel_adds_manual_review_sheet_without_changing_main_columns(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI产品", "北京", target_role="AI产品相关岗位",
                target_type="domain_scope",
            )
            for job_id, status in (("matched", "completed"), ("review", "manual_review")):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": job_id, "title": job_id,
                        "city_name": "北京", "job_link": f"https://example.test/{job_id}",
                    },
                )
                database.update_job(
                    task_id, job_id, crawl_status="completed", ai_status=status,
                    relevance_reason="职责边界不清" if status == "manual_review" else "目标一致",
                    relevance_confidence=0.55 if status == "manual_review" else 0.95,
                )

            workbook = load_workbook(export_task_to_excel(database, task_id, directory))

            self.assertEqual(workbook.sheetnames, ["岗位信息", "待人工确认"])
            self.assertEqual([cell.value for cell in workbook["岗位信息"][1]], EXCEL_COLUMNS)
            review = workbook["待人工确认"]
            self.assertEqual(review.max_row, 2)
            self.assertEqual(review["A2"].value, "review")
            self.assertEqual(review["D2"].value, "职责边界不清")
            self.assertEqual(review["E2"].value, 0.55)


class LoginManagerTests(unittest.TestCase):
    def test_login_manager_forwards_persistent_budget(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            budget = mock.Mock()
            manager = LoginManager(database, request_budget=budget)
            available = core.LoginProbeResult(core.LoginProbeStatus.AVAILABLE)

            with mock.patch.object(Path, "exists", return_value=True), \
                    mock.patch.object(
                        core, "require_runtime_dependencies", return_value=True
                    ), \
                    mock.patch.object(core, "is_cdp_ready", return_value=True), \
                    mock.patch.object(
                        core, "check_login_state", return_value=available
                    ) as check:
                state = manager.status()

            self.assertEqual(state.status, LoginStatus.LOGGED_IN)
            check.assert_called_once_with(
                core.DEFAULT_CDP_PORT, request_budget=budget
            )

    def test_login_status_classification_distinguishes_waiting_and_expired(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            chrome = root / "chrome.exe"
            profile = root / "profile"
            chrome.write_text("", encoding="utf-8")
            profile.mkdir()
            database = Database(root / "jobs.db")
            unauthenticated = core.LoginProbeResult(core.LoginProbeStatus.UNAUTHENTICATED)
            with mock.patch.object(core, "DEFAULT_CHROME_PATH", str(chrome)), \
                    mock.patch.object(core, "DEFAULT_CDP_DATA_DIR", str(profile)), \
                    mock.patch.object(core, "require_runtime_dependencies", return_value=True), \
                    mock.patch.object(core, "is_cdp_ready", return_value=True), \
                    mock.patch.object(core, "check_login_state", return_value=unauthenticated):
                manager = LoginManager(database)
                self.assertEqual(manager.status().status, LoginStatus.WAITING_FOR_LOGIN)
                database.set_state("last_login_success_at", "2026-08-01T00:00:00+00:00")
                self.assertEqual(manager.status().status, LoginStatus.EXPIRED)

    def test_restricted_login_is_not_reported_as_logged_out(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            chrome = root / "chrome.exe"
            profile = root / "profile"
            chrome.write_text("", encoding="utf-8")
            profile.mkdir()
            database = Database(root / "jobs.db")
            restricted = core.LoginProbeResult(core.LoginProbeStatus.RESTRICTED, code=37)
            with mock.patch.object(core, "DEFAULT_CHROME_PATH", str(chrome)), \
                    mock.patch.object(core, "DEFAULT_CDP_DATA_DIR", str(profile)), \
                    mock.patch.object(core, "require_runtime_dependencies", return_value=True), \
                    mock.patch.object(core, "is_cdp_ready", return_value=True), \
                    mock.patch.object(core, "check_login_state", return_value=restricted):
                self.assertEqual(LoginManager(database).status().status, LoginStatus.RESTRICTED)

    def test_recent_success_can_avoid_an_immediate_duplicate_probe(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            chrome = root / "chrome.exe"
            profile = root / "profile"
            chrome.write_text("", encoding="utf-8")
            profile.mkdir()
            database = Database(root / "jobs.db")
            database.set_state("last_login_success_at", utc_now())
            with mock.patch.object(core, "DEFAULT_CHROME_PATH", str(chrome)), \
                    mock.patch.object(core, "DEFAULT_CDP_DATA_DIR", str(profile)), \
                    mock.patch.object(core, "require_runtime_dependencies", return_value=True), \
                    mock.patch.object(core, "is_cdp_ready", return_value=True), \
                    mock.patch.object(core, "check_login_state") as probe:
                state = LoginManager(database).status(probe=True, allow_recent=True)
            self.assertEqual(state.status, LoginStatus.LOGGED_IN)
            probe.assert_not_called()


class TaskManagerTests(unittest.TestCase):
    def _strategy_task(self, database):
        spec = StrategySpec.create(
            "AI产品经理", "AI产品经理", "exact_role", ["上海"],
        )
        strategy = database.get_or_create_strategy(spec)
        return database.create_task(
            "AI产品经理",
            "上海",
            target_role="AI产品经理",
            target_type="exact_role",
            strategy_id=strategy["strategy_id"],
            scan_cycle=1,
            first_run_id="run-owned",
        )

    def test_start_reserves_worker_and_launches_module(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            process = mock.Mock(pid=4321)
            popen = mock.Mock(return_value=process)
            manager = TaskManager(database, popen=popen)
            self.assertTrue(manager.start(task_id))
            task = database.get_task(task_id)
            self.assertEqual(task["worker_pid"], 4321)
            command = popen.call_args.args[0]
            self.assertEqual(command[:3], [os.sys.executable, "-m", "boss_app.worker"])
            self.assertEqual(popen.call_args.kwargs["env"]["PYTHONIOENCODING"], "utf-8")

    def test_pause_sets_cooperative_flag(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            TaskManager(database, popen=mock.Mock()).pause(task_id)
            self.assertEqual(database.get_task(task_id)["pause_requested"], 1)

    def test_retry_ai_does_not_requeue_irrelevant_jobs(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "全国")
            database.upsert_job(task_id, {"job_id": "rejected", "title": "AI销售"})
            database.update_job(
                task_id, "rejected", crawl_status="completed", ai_status="irrelevant",
            )
            process = mock.Mock(pid=4321)
            manager = TaskManager(database, popen=mock.Mock(return_value=process))

            self.assertTrue(manager.retry_ai(task_id))

            self.assertEqual(database.list_jobs(task_id)[0]["ai_status"], "irrelevant")

    def test_generic_worker_controls_reject_strategy_tasks_without_mutation(self):
        for action_name in ("start", "resume", "expand", "retry_ai"):
            with self.subTest(action=action_name), tempfile.TemporaryDirectory() as directory:
                database = Database(Path(directory) / "jobs.db")
                task_id = self._strategy_task(database)
                database.upsert_job(task_id, {
                    "job_id": "saved",
                    "encrypt_job_id": "stable-saved",
                    "title": "岗位",
                    "job_link": "https://www.zhipin.com/job_detail/stable-saved.html",
                })
                database.update_job(
                    task_id, "saved", crawl_status="completed", ai_status="failed",
                )
                process = mock.Mock(pid=4321)
                manager = TaskManager(
                    database, popen=mock.Mock(return_value=process),
                )
                before_task = database.get_task(task_id)
                before_job = database.list_jobs(task_id)[0]
                actions = {
                    "start": lambda: manager.start(task_id),
                    "resume": lambda: manager.resume(task_id),
                    "expand": lambda: manager.expand(
                        task_id,
                        max_jobs=99,
                        max_pages=15,
                        run_mode="自定义数量",
                    ),
                    "retry_ai": lambda: manager.retry_ai(task_id),
                }

                action = actions[action_name]
                with self.assertRaisesRegex(RuntimeError, "Strategy Runner"):
                    action()

                after_task = database.get_task(task_id)
                after_job = database.list_jobs(task_id)[0]
                self.assertEqual(after_task["max_jobs"], before_task["max_jobs"])
                self.assertEqual(after_task["max_pages"], before_task["max_pages"])
                self.assertEqual(after_task["status"], before_task["status"])
                self.assertEqual(after_job["ai_status"], before_job["ai_status"])
                manager.popen.assert_not_called()


class _ConfiguredFailingParser:
    config = AIConfig(api_key="x", base_url="https://ai.test/v1", model="m")

    def __init__(self):
        self.calls = []

    def parse(self, full_jd, context=None):
        self.calls.append((full_jd, context))
        raise AIParseError("provider unavailable")


class _RelevanceParser:
    config = AIConfig(api_key="x", base_url="https://ai.test/v1", model="m")

    def __init__(self):
        self.calls = []

    def parse(self, full_jd, context=None):
        self.calls.append((full_jd, context))
        index = int((context or {})["job_name"].rsplit(" ", 1)[-1])
        return ParsedJD(
            is_ai_operations=index >= 10,
            job_type="全职",
            job_responsibilities=["负责 AI 运营"],
            job_requirements=["具备运营经验"],
            bonus_points=["无"],
        )


class _EventParser:
    config = AIConfig(api_key="x", base_url="https://ai.test/v1", model="m")

    def __init__(self, events, error=None):
        self.events = events
        self.error = error

    def parse(self, full_jd, context=None):
        job_name = (context or {})["job_name"]
        self.events.append(f"parse:{job_name}")
        if self.error:
            raise self.error
        return ParsedJD(
            is_ai_operations=True,
            job_type="全职",
            job_responsibilities=["负责 AI 产品运营"],
            job_requirements=["具备运营经验"],
            bonus_points=["无"],
        )


class _BlockingParser(_EventParser):
    def __init__(self, started, release, finished):
        super().__init__([])
        self.started = started
        self.release = release
        self.finished = finished

    def parse(self, full_jd, context=None):
        self.started.set()
        self.release.wait(2)
        self.finished.set()
        return super().parse(full_jd, context)


class _SlowFailingParser(_EventParser):
    def __init__(self, release):
        super().__init__([])
        self.release = release
        self.calls = 0

    def parse(self, full_jd, context=None):
        self.calls += 1
        self.release.wait(2)
        raise AIParseError("provider unavailable")


class CollectorTests(unittest.TestCase):
    @staticmethod
    def _core_mock():
        core_module = mock.Mock()
        core_module.DetailLoginRequiredError = core.DetailLoginRequiredError
        core_module.DetailExtractionError = core.DetailExtractionError
        core_module.is_access_restriction_error.return_value = False
        return core_module

    def test_detail_fetch_continues_while_previous_ai_is_running(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI产品运营", "北京", max_jobs=2)
            for index, name in enumerate(("岗位一", "岗位二"), 1):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": f"job-{index}", "title": name,
                        "job_link": f"https://example.test/job/{index}",
                    },
                )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            ai_started = threading.Event()
            release_ai = threading.Event()
            ai_finished = threading.Event()
            overlap_seen = []
            core_module = self._core_mock()

            def fetch(job, **_kwargs):
                if job["title"] == "岗位二":
                    first = database.list_jobs(task_id)[0]
                    overlap_seen.append(
                        first["ai_status"] == "processing" and not ai_finished.is_set()
                    )
                    release_ai.set()
                return {
                    "title": job["title"], "location": "北京", "salary": "20-30K",
                    "jd": f"{job['title']}完整 JD", "detail_url": job["job_link"],
                }

            core_module.fetch_job_detail.side_effect = fetch
            collector = Collector(
                database,
                ai_parser=_BlockingParser(ai_started, release_ai, ai_finished),
                core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            self.assertEqual(overlap_seen, [True])
            self.assertTrue(ai_finished.is_set())
            self.assertEqual(
                [row["ai_status"] for row in database.list_jobs(task_id)],
                ["completed", "completed"],
            )

    def test_detail_access_restriction_waits_for_saved_ai_backlog(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI产品运营", "北京", max_jobs=3)
            for index in (1, 2, 3):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": f"job-{index}", "title": f"岗位{index}",
                        "job_link": f"https://example.test/job/{index}",
                    },
                )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            ai_started = threading.Event()
            release_ai = threading.Event()
            ai_finished = threading.Event()
            restriction_overlapped_ai = []
            core_module = self._core_mock()

            def fetch(job, **_kwargs):
                if job["title"] == "岗位3":
                    restriction_overlapped_ai.append(
                        ai_started.is_set() and not ai_finished.is_set()
                    )
                    release_ai.set()
                    raise RuntimeError("BOSS access restricted: code: 37")
                return {
                    "title": job["title"], "location": "北京", "salary": "20-30K",
                    "jd": f"{job['title']}完整 JD", "detail_url": job["job_link"],
                }

            core_module.fetch_job_detail.side_effect = fetch
            core_module.is_access_restriction_error.side_effect = (
                lambda error: "code: 37" in str(error)
            )
            collector = Collector(
                database,
                ai_parser=_BlockingParser(ai_started, release_ai, ai_finished),
                core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            rows = database.list_jobs(task_id)
            self.assertEqual(restriction_overlapped_ai, [True])
            self.assertEqual(database.get_task(task_id)["status"], "waiting_for_access")
            self.assertEqual([row["ai_status"] for row in rows[:2]], ["completed", "completed"])
            self.assertEqual(rows[2]["crawl_status"], "pending")

    def test_restricted_preflight_drains_saved_complete_jd_before_stopping(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI产品运营", "北京")
            database.upsert_job(
                task_id,
                {
                    "job_id": "saved", "title": "已保存岗位",
                    "job_link": "https://example.test/job/saved",
                },
            )
            database.update_job(
                task_id, "saved", crawl_status="completed", full_jd="已保存完整 JD",
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            parser = _EventParser([])
            collector = Collector(database, ai_parser=parser, core_module=self._core_mock())

            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.RESTRICTED, "code: 37"),
            ):
                collector.run(task_id, token)

            self.assertEqual(database.list_jobs(task_id)[0]["ai_status"], "completed")
            self.assertEqual(database.get_task(task_id)["status"], "waiting_for_access")

    def test_saved_ai_backlog_runs_before_new_detail_requests(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI产品运营", "上海", max_jobs=2)
            database.upsert_job(
                task_id,
                {
                    "job_id": "history", "title": "历史岗位",
                    "job_link": "https://example.test/job/history",
                },
            )
            database.update_job(
                task_id, "history", crawl_status="completed", full_jd="历史完整 JD",
            )
            database.upsert_job(
                task_id,
                {
                    "job_id": "new", "title": "新增岗位",
                    "job_link": "https://example.test/job/new",
                },
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            events = []
            core_module = self._core_mock()

            def fetch(job, **_kwargs):
                events.append(f"fetch:{job['title']}")
                return {
                    "title": job["title"], "location": "上海", "salary": "20-30K",
                    "jd": "新增完整 JD", "detail_url": job["job_link"],
                }

            core_module.fetch_job_detail.side_effect = fetch
            collector = Collector(
                database, ai_parser=_EventParser(events), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            self.assertEqual(events[0], "parse:历史岗位")
            self.assertEqual(events[1:], ["fetch:新增岗位", "parse:新增岗位"])

    def test_ai_failure_stops_before_next_detail_request(self):
        parsers = (
            JDParser(AIConfig()),
            _EventParser([], AIParseError("provider unavailable")),
        )
        for parser in parsers:
            with self.subTest(parser=type(parser).__name__), tempfile.TemporaryDirectory() as directory:
                database = Database(Path(directory) / "jobs.db")
                task_id = database.create_task("AI产品运营", "北京", max_jobs=3)
                for index in (1, 2, 3):
                    database.upsert_job(
                        task_id,
                        {
                            "job_id": f"job-{index}", "title": f"岗位{index}",
                            "job_link": f"https://example.test/job/{index}",
                        },
                    )
                token = "worker"
                self.assertTrue(database.reserve_worker(task_id, token))
                core_module = self._core_mock()
                core_module.fetch_job_detail.side_effect = lambda job, **_kwargs: {
                    "title": job["title"], "location": "北京", "salary": "20-30K",
                    "jd": "完整 JD", "detail_url": job["job_link"],
                }
                collector = Collector(
                    database, ai_parser=parser, core_module=core_module,
                    sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
                )
                with mock.patch(
                    "boss_app.collector.LoginManager.status",
                    return_value=LoginState(LoginStatus.LOGGED_IN),
                ):
                    collector.run(task_id, token)

                rows = database.list_jobs(task_id)
                self.assertEqual(database.get_task(task_id)["status"], "waiting_for_ai")
                self.assertIn(core_module.fetch_job_detail.call_count, (1, 2))
                self.assertEqual(rows[0]["crawl_status"], "completed")
                self.assertEqual(rows[0]["ai_status"], "waiting_for_ai")
                self.assertEqual(rows[2]["crawl_status"], "pending")

    def test_slow_ai_failure_does_not_run_already_queued_ai_requests(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI产品运营", "北京", max_jobs=3)
            for index in (1, 2, 3):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": f"job-{index}", "title": f"岗位{index}",
                        "job_link": f"https://example.test/job/{index}",
                    },
                )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            release = threading.Event()
            parser = _SlowFailingParser(release)
            core_module = self._core_mock()

            def fetch(job, **_kwargs):
                if job["title"] == "岗位3":
                    release.set()
                return {
                    "title": job["title"], "location": "北京", "salary": "20-30K",
                    "jd": "完整 JD", "detail_url": job["job_link"],
                }

            core_module.fetch_job_detail.side_effect = fetch
            collector = Collector(
                database, ai_parser=parser, core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            self.assertEqual(core_module.fetch_job_detail.call_count, 3)
            self.assertEqual(parser.calls, 1)
            self.assertEqual(database.get_task(task_id)["status"], "waiting_for_ai")

    def test_city_ai_product_irrelevant_is_excluded_without_replenishment_target(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI产品运营", "北京", max_pages=15, max_jobs=450,
            )
            database.upsert_job(
                task_id,
                {
                    "job_id": "irrelevant", "title": "AI产品运营 0",
                    "job_link": "https://example.test/job/irrelevant",
                },
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()
            core_module.scrape_list.return_value = {"total": 0, "jobs": []}
            core_module.fetch_job_detail.return_value = {
                "title": "AI产品运营 0", "location": "北京", "salary": "20-30K",
                "jd": "普通产品经理岗位描述", "detail_url": "https://example.test/job/irrelevant",
            }
            collector = Collector(
                database, ai_parser=_RelevanceParser(), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            self.assertEqual(database.list_jobs(task_id)[0]["ai_status"], "irrelevant")
            self.assertEqual(database.get_task(task_id)["status"], "completed")
            self.assertEqual(core_module.scrape_list.call_count, 1)

    def test_city_expansion_forwards_fifteen_pages_and_remaining_candidates(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI产品运营", "北京", max_pages=15, max_jobs=450,
            )
            expected_keys = set()
            for index in range(20):
                job_id = f"saved-{index}"
                job_url = f"https://example.test/job/{index}"
                expected_keys.add(job_url)
                database.upsert_job(
                    task_id,
                    {"job_id": job_id, "title": f"岗位 {index}", "job_link": job_url},
                )
                database.update_job(
                    task_id, job_id, crawl_status="completed", ai_status="completed",
                )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()
            collector = Collector(database, core_module=core_module)

            self.assertTrue(collector._collect_list(database.get_task(task_id), token))

            call = core_module.scrape_list.call_args
            self.assertEqual(call.args[2], 15)
            self.assertEqual(call.kwargs["max_jobs"], 430)
            self.assertEqual(call.kwargs["existing_keys"], expected_keys)

    def test_beijing_expansion_does_not_reprocess_saved_completed_jobs(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI产品运营", "北京", max_pages=15, max_jobs=450,
            )
            for index in range(20):
                job_id = f"saved-{index}"
                database.upsert_job(
                    task_id,
                    {
                        "job_id": job_id, "title": f"已完成岗位 {index}",
                        "job_link": f"https://example.test/saved/{index}",
                    },
                )
                database.update_job(
                    task_id, job_id, crawl_status="completed", ai_status="completed",
                    full_jd="已保存完整 JD",
                )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            events = []
            core_module = self._core_mock()

            def emit_new(*_args, **kwargs):
                kwargs["on_job"]({
                    "job_id": "new", "title": "新增岗位",
                    "job_link": "https://example.test/new",
                })
                return {"total": 1, "jobs": []}

            core_module.scrape_list.side_effect = emit_new

            def fetch(job, **_kwargs):
                events.append(f"fetch:{job['title']}")
                return {
                    "title": job["title"], "location": "北京", "salary": "20-30K",
                    "jd": "新增完整 JD", "detail_url": job["job_link"],
                }

            core_module.fetch_job_detail.side_effect = fetch
            collector = Collector(
                database, ai_parser=_EventParser(events), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            self.assertEqual(events, ["fetch:新增岗位", "parse:新增岗位"])
            self.assertEqual(core_module.fetch_job_detail.call_count, 1)

    def test_shanghai_saved_ai_backlog_precedes_first_new_boss_request(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI产品运营", "上海", max_pages=15, max_jobs=450,
            )
            for index in range(20):
                job_id = f"saved-{index}"
                database.upsert_job(
                    task_id,
                    {
                        "job_id": job_id, "title": f"历史岗位 {index}",
                        "job_link": f"https://example.test/saved/{index}",
                    },
                )
                database.update_job(
                    task_id, job_id, crawl_status="completed",
                    ai_status="completed" if index < 8 else "pending",
                    full_jd="历史完整 JD",
                )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            events = []
            core_module = self._core_mock()

            def list_jobs(*_args, **_kwargs):
                events.append("list")
                return {"total": 0, "jobs": []}

            core_module.scrape_list.side_effect = list_jobs
            collector = Collector(
                database, ai_parser=_EventParser(events), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            self.assertEqual(
                events[:12], [f"parse:历史岗位 {index}" for index in range(8, 20)],
            )
            self.assertEqual(events[12:], ["list"])

    def test_national_ai_operations_replenishes_rejected_candidates_to_fifty(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI运营", "全国", max_pages=10, max_jobs=50,
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()
            next_index = 0

            def emit_jobs(*_args, **kwargs):
                nonlocal next_index
                count = kwargs["max_jobs"]
                for _ in range(count):
                    index = next_index
                    next_index += 1
                    kwargs["on_job"]({
                        "job_id": f"job-{index}",
                        "title": f"AI运营 {index}",
                        "city_name": "上海",
                        "location": "上海·浦东新区",
                        "job_link": f"https://example.test/job/{index}",
                    })
                return {"total": count, "jobs": []}

            core_module.scrape_list.side_effect = emit_jobs
            core_module.fetch_job_detail.side_effect = lambda job, **_kwargs: {
                "title": job["title"],
                "location": "上海·浦东新区·张江",
                "salary": "20-30K",
                "jd": "完整岗位描述：负责 AI 运营工作",
                "detail_url": job["job_link"],
            }
            collector = Collector(
                database, ai_parser=_RelevanceParser(), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            logged_in = LoginState(LoginStatus.LOGGED_IN)

            with mock.patch(
                "boss_app.collector.LoginManager.status", return_value=logged_in,
            ):
                collector.run(task_id, token)

            snapshot = database.snapshot(task_id)
            self.assertEqual(core_module.scrape_list.call_count, 2)
            self.assertEqual(len(database.list_jobs(task_id)), 60)
            self.assertEqual(snapshot["qualified"], 50)
            self.assertEqual(snapshot["irrelevant"], 10)
            self.assertEqual(database.get_task(task_id)["status"], "completed")
            self.assertEqual(
                core_module.scrape_list.call_args_list[1].kwargs["max_jobs"], 10,
            )
            self.assertTrue(all(
                job["city"] == "上海" for job in database.list_jobs(task_id)
            ))

    def test_national_ai_operations_without_new_candidates_is_incomplete(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI运营", "全国", max_pages=10, max_jobs=50,
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()
            core_module.scrape_list.return_value = {"total": 0, "jobs": []}
            collector = Collector(
                database, ai_parser=_RelevanceParser(), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )
            logged_in = LoginState(LoginStatus.LOGGED_IN)

            with mock.patch(
                "boss_app.collector.LoginManager.status", return_value=logged_in,
            ):
                collector.run(task_id, token)

            self.assertEqual(database.get_task(task_id)["status"], "incomplete")
            self.assertEqual(core_module.scrape_list.call_count, 1)

    def test_national_resume_processes_saved_candidate_before_requesting_more(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI运营", "全国", max_pages=10, max_jobs=1,
            )
            database.upsert_job(
                task_id,
                {
                    "job_id": "saved",
                    "title": "AI运营 10",
                    "city_name": "上海",
                    "job_link": "https://example.test/job/saved",
                },
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()
            core_module.fetch_job_detail.return_value = {
                "title": "AI运营 10", "location": "上海·浦东新区",
                "salary": "20-30K", "jd": "完整岗位描述：负责 AI 运营工作",
                "detail_url": "https://example.test/job/saved",
            }
            collector = Collector(
                database, ai_parser=_RelevanceParser(), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )

            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            self.assertEqual(database.get_task(task_id)["status"], "completed")
            self.assertEqual(database.snapshot(task_id)["qualified"], 1)
            core_module.scrape_list.assert_not_called()

    def test_existing_only_processes_all_saved_candidates_without_listing(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI运营", "全国", max_pages=15, max_jobs=50,
            )
            for index in (10, 11):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": f"saved-{index}",
                        "title": f"AI运营 {index}",
                        "job_link": f"https://example.test/job/{index}",
                    },
                )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()
            core_module.fetch_job_detail.side_effect = lambda job, **_kwargs: {
                "title": job["title"], "location": "上海·浦东新区",
                "salary": "20-30K", "jd": "完整岗位描述：负责 AI 运营工作",
                "detail_url": job["job_link"],
            }
            collector = Collector(
                database, ai_parser=_RelevanceParser(), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )

            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run_existing(task_id, token)

            self.assertEqual(core_module.fetch_job_detail.call_count, 2)
            core_module.scrape_list.assert_not_called()
            self.assertEqual(database.snapshot(task_id)["qualified"], 2)
            self.assertEqual(database.get_task(task_id)["status"], "completed")

    def test_existing_only_resumes_after_detail_restriction_without_listing(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI运营", "全国", max_pages=15, max_jobs=50,
            )
            database.upsert_job(
                task_id,
                {
                    "job_id": "saved-10", "title": "AI运营 10",
                    "job_link": "https://example.test/job/10",
                },
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()
            restricted = core.AccessRestrictedError(
                "BOSS access restricted: code: 37",
            )
            core_module.fetch_job_detail.side_effect = [
                restricted,
                {
                    "title": "AI运营 10", "location": "上海·浦东新区",
                    "salary": "20-30K", "jd": "完整岗位描述：负责 AI 运营工作",
                    "detail_url": "https://example.test/job/10",
                },
            ]
            core_module.is_access_restriction_error.side_effect = (
                core.is_access_restriction_error
            )
            collector = Collector(
                database, ai_parser=_RelevanceParser(), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )

            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run_existing(task_id, token)
                self.assertEqual(
                    database.get_task(task_id)["status"], "waiting_for_access",
                )
                self.assertEqual(
                    database.list_jobs(task_id)[0]["crawl_attempts"], 0,
                )
                collector.run_existing(task_id, token)

            core_module.scrape_list.assert_not_called()
            self.assertEqual(database.snapshot(task_id)["qualified"], 1)
            self.assertEqual(database.get_task(task_id)["status"], "completed")

    def test_national_ai_operations_stops_replenishing_while_ai_is_unavailable(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task(
                "AI运营", "全国", max_pages=10, max_jobs=50,
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()

            def emit_job(*_args, **kwargs):
                kwargs["on_job"]({
                    "job_id": "job-0",
                    "title": "AI运营 0",
                    "city_name": "上海",
                    "job_link": "https://example.test/job/0",
                })
                return {"total": 1, "jobs": []}

            core_module.scrape_list.side_effect = emit_job
            core_module.fetch_job_detail.return_value = {
                "title": "AI运营 0", "location": "上海·浦东新区",
                "salary": "20-30K", "jd": "完整岗位描述",
                "detail_url": "https://example.test/job/0",
            }
            collector = Collector(
                database, ai_parser=_ConfiguredFailingParser(), core_module=core_module,
                sleep_fn=lambda _seconds: None, jitter_fn=lambda *_args: 0,
            )

            with mock.patch(
                "boss_app.collector.LoginManager.status",
                return_value=LoginState(LoginStatus.LOGGED_IN),
            ):
                collector.run(task_id, token)

            self.assertEqual(database.get_task(task_id)["status"], "waiting_for_ai")
            self.assertEqual(core_module.scrape_list.call_count, 1)
            self.assertEqual(database.list_jobs(task_id)[0]["ai_status"], "waiting_for_ai")

    def test_non_national_tasks_keep_legacy_completion_semantics(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            database.upsert_job(
                task_id, {"job_id": "legacy", "title": "AI运营 0"},
            )
            database.update_job(
                task_id, "legacy", crawl_status="completed", full_jd="完整 JD",
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            parser = _RelevanceParser()
            collector = Collector(database, ai_parser=parser)

            self.assertTrue(collector._process_ai(task_id, token))

            self.assertEqual(database.list_jobs(task_id)[0]["ai_status"], "completed")

    def test_local_scheduler_expands_existing_ten_to_fifty(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳", max_jobs=10)
            for index in range(10):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": f"saved-{index}",
                        "title": f"已完成 {index}",
                        "job_link": f"https://example.test/saved/{index}",
                    },
                )
                database.update_job(task_id, f"saved-{index}", crawl_status="completed")
            database.update_task_limits(
                task_id, max_jobs=50, max_pages=3, run_mode="50条扩容测试",
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = mock.Mock()

            def emit_jobs(*_args, **kwargs):
                self.assertEqual(kwargs["max_jobs"], 40)
                for index in range(10, 50):
                    kwargs["on_job"]({
                        "job_id": f"new-{index}",
                        "title": f"新增 {index}",
                        "job_link": f"https://example.test/new/{index}",
                    })
                return {"jobs": []}

            core_module.scrape_list.side_effect = emit_jobs
            collector = Collector(database, core_module=core_module)

            self.assertTrue(collector._collect_list(database.get_task(task_id), token))

            rows = database.list_jobs(task_id)
            self.assertEqual(len(rows), 50)
            self.assertEqual(
                sum(row["crawl_status"] == "completed" for row in rows), 10,
            )

    def test_restricted_login_marks_task_waiting_for_access(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "全国")
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            collector = Collector(database)
            restricted = LoginState(LoginStatus.RESTRICTED, "code: 37")

            with mock.patch("boss_app.collector.LoginManager.status", return_value=restricted):
                collector.run(task_id, token)

            task = database.get_task(task_id)
            self.assertEqual(task["status"], "waiting_for_access")
            self.assertIn("code: 37", task["error_message"])

    def test_access_restriction_during_list_stops_without_retry_loop(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳", max_jobs=50)
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = mock.Mock()
            restricted = core.AccessRestrictedError("BOSS access restricted: code: 37")
            core_module.scrape_list.side_effect = restricted
            core_module.is_access_restriction_error.side_effect = core.is_access_restriction_error
            collector = Collector(database, core_module=core_module)

            result = collector._collect_list(database.get_task(task_id), token)

            self.assertFalse(result)
            self.assertEqual(core_module.scrape_list.call_count, 1)
            self.assertEqual(database.get_task(task_id)["status"], "waiting_for_access")
            self.assertEqual(database.list_jobs(task_id), [])

    def test_access_restriction_does_not_advance_list_page_cursor(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳", max_pages=10, max_jobs=50)
            database.update_task(task_id, list_next_page=6)
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = mock.Mock()
            restricted = core.AccessRestrictedError("BOSS access restricted: code: 37")
            core_module.scrape_list.side_effect = restricted
            core_module.is_access_restriction_error.side_effect = core.is_access_restriction_error
            collector = Collector(database, core_module=core_module)

            result = collector._collect_list(database.get_task(task_id), token)

            self.assertFalse(result)
            self.assertEqual(core_module.scrape_list.call_args.kwargs["start_page"], 6)
            self.assertEqual(database.get_task(task_id)["list_next_page"], 6)
            self.assertEqual(database.get_task(task_id)["status"], "waiting_for_access")

    def test_access_restriction_during_details_stops_without_retry_loop(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            for job_id in ("one", "two"):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": job_id,
                        "title": f"岗位 {job_id}",
                        "job_link": f"https://example.test/{job_id}",
                    },
                )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            collector = Collector(database)
            restricted = RuntimeError("BOSS 接口返回限制状态（code: 37）")

            with mock.patch.object(core, "fetch_job_detail", side_effect=restricted) as fetch:
                result = collector._collect_details(task_id, token)

            self.assertFalse(result)
            self.assertEqual(fetch.call_count, 1)
            self.assertEqual(database.get_task(task_id)["status"], "waiting_for_access")
            rows = database.list_jobs(task_id)
            self.assertTrue(all(row["crawl_status"] == "pending" for row in rows))
            self.assertTrue(all(row["crawl_attempts"] == 0 for row in rows))

    def test_completed_and_exhausted_failed_jobs_are_not_reprocessed(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            for job_id in ("completed", "exhausted", "pending"):
                database.upsert_job(
                    task_id,
                    {
                        "job_id": job_id,
                        "title": job_id,
                        "job_link": f"https://example.test/{job_id}",
                    },
                )
            database.update_job(task_id, "completed", crawl_status="completed")
            database.update_job(task_id, "exhausted", crawl_status="failed", crawl_attempts=2)
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            collector = Collector(database, ai_parser=_EventParser([]))
            detail = {
                "title": "pending", "location": "深圳", "salary": "10-20K",
                "jd": "完整岗位描述", "detail_url": "https://example.test/pending",
            }

            with mock.patch.object(core, "fetch_job_detail", return_value=detail) as fetch:
                self.assertTrue(collector._collect_details(task_id, token))

            self.assertEqual(fetch.call_count, 1)
            self.assertEqual(fetch.call_args.args[0]["job_id"], "pending")
            self.assertEqual(database.list_jobs(task_id)[1]["crawl_status"], "failed")

    def test_network_pause_uses_configured_random_jitter(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            collector = Collector(database)
            pause = getattr(collector, "_network_pause", None)
            self.assertIsNotNone(pause, "缺少串行网络访问间隔")
            collector.network_interval_min = 2.0
            collector.network_interval_max = 5.0
            collector.sleep_fn = mock.Mock()
            collector.jitter_fn = mock.Mock(return_value=3.5)

            pause()

            collector.jitter_fn.assert_called_once_with(2.0, 5.0)
            collector.sleep_fn.assert_called_once_with(3.5)

    def test_resumed_list_passes_existing_job_keys_to_core(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳", max_jobs=10)
            database.upsert_job(
                task_id,
                {
                    "job_id": "existing",
                    "title": "AI运营专员",
                    "job_link": "https://example.test/job/existing",
                },
            )
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = mock.Mock()
            collector = Collector(database, core_module=core_module)

            self.assertTrue(collector._collect_list(database.get_task(task_id), token))

            self.assertEqual(
                core_module.scrape_list.call_args.kwargs["existing_keys"],
                {"https://example.test/job/existing"},
            )
            self.assertEqual(core_module.scrape_list.call_args.kwargs["max_jobs"], 9)

    def test_resumed_list_uses_and_persists_page_cursor(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳", max_pages=10, max_jobs=50)
            database.update_task(task_id, list_next_page=6)
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            core_module = self._core_mock()

            def complete_page(*_args, **kwargs):
                self.assertEqual(kwargs["start_page"], 6)
                kwargs["on_page_complete"](7)
                return {"jobs": []}

            core_module.scrape_list.side_effect = complete_page
            collector = Collector(database, core_module=core_module)

            self.assertTrue(collector._collect_list(database.get_task(task_id), token))

            self.assertEqual(database.get_task(task_id)["list_next_page"], 7)

    def test_ai_failure_marks_waiting_without_dropping_job(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            database.upsert_job(task_id, {"job_id": "abc", "title": "AI运营"})
            database.update_job(task_id, "abc", crawl_status="completed", full_jd="完整 JD")
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            parser = _ConfiguredFailingParser()
            collector = Collector(database, ai_parser=parser)
            self.assertFalse(collector._process_ai(task_id, token))
            row = database.list_jobs(task_id)[0]
            self.assertEqual(row["ai_status"], "waiting_for_ai")
            self.assertEqual(row["full_jd"], "完整 JD")
            self.assertEqual(row["job_responsibilities"], "待处理")

    def test_empty_jd_is_marked_invalid_without_calling_ai(self):
        with tempfile.TemporaryDirectory() as directory:
            database = Database(Path(directory) / "jobs.db")
            task_id = database.create_task("AI运营", "深圳")
            database.upsert_job(task_id, {"job_id": "abc", "title": "AI运营"})
            database.update_job(task_id, "abc", crawl_status="completed", full_jd="")
            token = "worker"
            self.assertTrue(database.reserve_worker(task_id, token))
            parser = _ConfiguredFailingParser()
            collector = Collector(database, ai_parser=parser)
            collector._process_ai(task_id, token)
            row = database.list_jobs(task_id)[0]
            self.assertEqual(row["ai_status"], "invalid")
            self.assertEqual(row["job_responsibilities"], "待处理")
            self.assertEqual(parser.calls, [])


if __name__ == "__main__":
    unittest.main()
