import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from boss_app.db import Database
from boss_app.exporter import (
    EXCEL_COLUMNS,
    export_strategy_run,
    freeze_strategy_run_snapshot,
)
from boss_app.strategy_model import StrategySpec


class StrategyExporterTests(unittest.TestCase):
    def setUp(self):
        self.tempdir = tempfile.TemporaryDirectory()
        self.directory = Path(self.tempdir.name)
        self.database = Database(self.directory / "jobs.db")
        self.spec = StrategySpec.create(
            "新媒体运营", "新媒体运营", "exact_role", ["北京"],
        )
        self.strategy = self.database.get_or_create_strategy(self.spec)
        self.run_one, _ = self.database.create_or_resume_run(
            self.strategy["strategy_id"], 1,
        )
        self.task_id = self.database.ensure_strategy_tasks(
            self.strategy["strategy_id"], 1, self.spec,
            first_run_id=self.run_one["run_id"],
        )[0]
        self.database.update_task(
            self.task_id, last_run_id=self.run_one["run_id"],
        )

    def tearDown(self):
        self.tempdir.cleanup()

    def _save_job(self, suffix):
        self.database.upsert_job(self.task_id, {
            "job_id": suffix,
            "encrypt_job_id": f"source-{suffix}",
            "title": f"岗位{suffix}",
            "city_name": "北京",
            "job_link": f"https://www.zhipin.com/job_detail/source-{suffix}.html",
        })
        row = self.database.list_jobs(self.task_id)[-1]
        self.database.update_job(
            self.task_id, row["job_id"],
            full_jd=f"完整 JD {suffix}",
            crawl_status="completed",
            ai_status="completed",
            job_responsibilities=[f"职责{suffix}"],
            job_requirements=[f"要求{suffix}"],
            bonus_points=["无"],
        )

    def test_running_run_cannot_freeze_or_export(self):
        with self.assertRaisesRegex(RuntimeError, "running"):
            freeze_strategy_run_snapshot(
                self.database,
                self.strategy["strategy_id"],
                self.run_one["run_id"],
            )

        with self.assertRaisesRegex(RuntimeError, "running"):
            export_strategy_run(
                self.database,
                self.strategy["strategy_id"],
                self.run_one["run_id"],
                self.directory,
            )

    def test_each_run_gets_an_independent_cumulative_snapshot(self):
        self._save_job("一")
        first = export_strategy_run(
            self.database,
            self.strategy["strategy_id"],
            self.run_one["run_id"],
            self.directory,
        )
        self.database.update_run(self.run_one["run_id"], status="completed")
        run_two, _ = self.database.create_or_resume_run(
            self.strategy["strategy_id"], 1,
        )
        self.database.update_task(self.task_id, last_run_id=run_two["run_id"])
        self._save_job("二")

        second = export_strategy_run(
            self.database,
            self.strategy["strategy_id"],
            run_two["run_id"],
            self.directory,
        )

        self.assertIn("Run001_累计1条", first.name)
        self.assertIn("Run002_累计2条", second.name)
        self.assertNotEqual(first, second)
        self.assertEqual(load_workbook(first)["岗位信息"].max_row - 1, 1)
        self.assertEqual(load_workbook(second)["岗位信息"].max_row - 1, 2)

    def test_zero_qualified_jobs_still_produces_a_valid_header_only_file(self):
        output = export_strategy_run(
            self.database,
            self.strategy["strategy_id"],
            self.run_one["run_id"],
            self.directory,
        )
        sheet = load_workbook(output)["岗位信息"]

        self.assertEqual([cell.value for cell in sheet[1]], EXCEL_COLUMNS)
        self.assertEqual(sheet.max_row, 1)
        self.assertIn("累计0条", output.name)

    def test_reexport_old_run_uses_its_frozen_snapshot_after_newer_run(self):
        self._save_job("一")
        original = export_strategy_run(
            self.database,
            self.strategy["strategy_id"],
            self.run_one["run_id"],
            self.directory,
        )
        self.assertEqual(
            export_strategy_run(
                self.database,
                self.strategy["strategy_id"],
                self.run_one["run_id"],
                self.directory,
            ),
            original,
        )
        self.database.update_run(self.run_one["run_id"], status="completed")
        run_two, _ = self.database.create_or_resume_run(
            self.strategy["strategy_id"], 1,
        )
        self.database.update_task(self.task_id, last_run_id=run_two["run_id"])
        self._save_job("二")
        latest = export_strategy_run(
            self.database,
            self.strategy["strategy_id"],
            run_two["run_id"],
            self.directory,
        )
        original.unlink()

        recovered = export_strategy_run(
            self.database,
            self.strategy["strategy_id"],
            self.run_one["run_id"],
            self.directory,
        )

        self.assertIn("Run001_累计1条", recovered.name)
        self.assertEqual(
            load_workbook(recovered)["岗位信息"].max_row - 1, 1,
        )
        self.assertEqual(
            self.database.get_strategy(self.strategy["strategy_id"])[
                "latest_output_path"
            ],
            str(latest),
        )

        self.database.update_strategy(
            self.strategy["strategy_id"], latest_output_path=str(recovered),
        )
        self.assertEqual(
            export_strategy_run(
                self.database,
                self.strategy["strategy_id"],
                run_two["run_id"],
                self.directory,
            ),
            latest,
        )
        self.assertEqual(
            self.database.get_strategy(self.strategy["strategy_id"])[
                "latest_output_path"
            ],
            str(latest),
        )

    def test_confirmed_unavailable_and_pending_ai_are_excluded(self):
        self._save_job("有效")
        self._save_job("失效")
        rows = self.database.list_jobs(self.task_id)
        self.database.update_job(
            self.task_id, rows[-1]["job_id"],
            availability_status="confirmed_unavailable",
            confirmed_unavailable_at="2026-08-09T00:00:00+00:00",
        )
        self.database.upsert_job(self.task_id, {
            "job_id": "pending",
            "encrypt_job_id": "source-pending",
            "title": "待处理",
            "job_link": "https://www.zhipin.com/job_detail/source-pending.html",
        })

        output = export_strategy_run(
            self.database,
            self.strategy["strategy_id"],
            self.run_one["run_id"],
            self.directory,
        )

        self.assertEqual(load_workbook(output)["岗位信息"].max_row - 1, 1)

    def test_validation_failure_keeps_database_and_marks_export_failed(self):
        self._save_job("一")

        with mock.patch(
            "boss_app.exporter.load_workbook",
            side_effect=InvalidFileException("cannot verify"),
        ):
            with self.assertRaisesRegex(InvalidFileException, "cannot verify"):
                export_strategy_run(
                    self.database,
                    self.strategy["strategy_id"],
                    self.run_one["run_id"],
                    self.directory,
                )

        run = self.database.get_run(self.run_one["run_id"])
        self.assertEqual(run["export_status"], "failed")
        self.assertEqual(self.database.count_catalog_jobs(), 1)
        self.assertEqual(list(self.directory.glob("*.tmp.xlsx")), [])

    def test_cleanup_failure_does_not_mask_the_original_export_error(self):
        self._save_job("一")

        with (
            mock.patch(
                "boss_app.exporter._write_workbook",
                side_effect=OSError("write failed"),
            ),
            mock.patch.object(
                Path, "unlink", side_effect=PermissionError("cleanup failed"),
            ) as unlink,
        ):
            with self.assertRaisesRegex(OSError, "write failed"):
                export_strategy_run(
                    self.database,
                    self.strategy["strategy_id"],
                    self.run_one["run_id"],
                    self.directory,
                )
            unlink.assert_called_once()

        run = self.database.get_run(self.run_one["run_id"])
        self.assertEqual(run["export_status"], "failed")
        self.assertEqual(run["export_error"], "write failed")

    def test_cleanup_failure_does_not_mask_keyboard_interrupt(self):
        with (
            mock.patch(
                "boss_app.exporter._write_workbook",
                side_effect=KeyboardInterrupt("stop"),
            ),
            mock.patch.object(
                Path, "unlink", side_effect=PermissionError("cleanup failed"),
            ) as unlink,
        ):
            with self.assertRaisesRegex(KeyboardInterrupt, "stop"):
                export_strategy_run(
                    self.database,
                    self.strategy["strategy_id"],
                    self.run_one["run_id"],
                    self.directory,
                )
            unlink.assert_called_once()

    def test_successful_publish_does_not_clean_the_replaced_temp_path(self):
        with mock.patch.object(
            Path, "unlink", side_effect=PermissionError("cleanup failed"),
        ) as unlink:
            output = export_strategy_run(
                self.database,
                self.strategy["strategy_id"],
                self.run_one["run_id"],
                self.directory,
            )

        self.assertTrue(output.is_file())
        unlink.assert_not_called()

    def test_programming_error_is_not_recorded_as_export_failure(self):
        with mock.patch(
            "boss_app.exporter._write_workbook",
            side_effect=AttributeError("programming bug"),
        ):
            with self.assertRaisesRegex(AttributeError, "programming bug"):
                export_strategy_run(
                    self.database,
                    self.strategy["strategy_id"],
                    self.run_one["run_id"],
                    self.directory,
                )

        run = self.database.get_run(self.run_one["run_id"])
        self.assertEqual(run["export_status"], "pending")
        self.assertEqual(run["export_error"], "")

    def test_output_directory_creation_failure_marks_export_failed(self):
        self._save_job("一")
        invalid_output_dir = self.directory / "not-a-directory"
        invalid_output_dir.write_text("occupied", encoding="utf-8")

        with self.assertRaises(OSError):
            export_strategy_run(
                self.database,
                self.strategy["strategy_id"],
                self.run_one["run_id"],
                invalid_output_dir,
            )

        run = self.database.get_run(self.run_one["run_id"])
        self.assertEqual(run["export_status"], "failed")
        self.assertTrue(run["export_error"])
        self.assertEqual(self.database.count_catalog_jobs(), 1)


if __name__ == "__main__":
    unittest.main()
